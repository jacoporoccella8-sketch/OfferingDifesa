"""
Defence Market Daily Scraper - NTT DATA Italia
"""

import os, sys, json, time, smtplib, logging
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
from anthropic import Anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

load_dotenv()

MODEL = "claude-sonnet-4-6"
PLAYERS_PER_RUN = 3
PAUSE_BETWEEN_PLAYERS = 10
MAX_RETRIES = 3
BACKOFF_SECONDS = [20, 40, 80]
WEB_SEARCH_MAX_USES = 4
DATA_DIR = "data"
OUTPUT_DIR = "output"
HISTORY_PATH = os.path.join(DATA_DIR, "history.json")
ROTATION_STATE_PATH = os.path.join(DATA_DIR, "rotation_state.json")

PLAYERS = [
    {"name": "Ministero della Difesa", "market": "Pubblico", "complexity": "Molto alta"},
    {"name": "Forze Armate (Esercito/Marina/AM)", "market": "Pubblico", "complexity": "Alta"},
    {"name": "Difesa Servizi", "market": "Pubblico", "complexity": "Media"},
    {"name": "Segretariato Generale Difesa/DNA", "market": "Pubblico", "complexity": "Alta"},
    {"name": "Leonardo", "market": "Privato", "complexity": "Alta"},
    {"name": "Fincantieri", "market": "Privato", "complexity": "Alta"},
    {"name": "MBDA Italia", "market": "Privato", "complexity": "Alta"},
    {"name": "Elettronica Group", "market": "Privato", "complexity": "Alta"},
    {"name": "Thales Alenia Space Italia", "market": "Privato", "complexity": "Alta"},
    {"name": "Avio Aero", "market": "Privato", "complexity": "Alta"},
    {"name": "Leonardo DRS", "market": "Privato", "complexity": "Alta"},
]

COMPETITORS = [
    {"name": "Leonardo",           "cluster": "Leader integrato",     "aliases": ["leonardo s.p.a", "leonardo spa", "finmeccanica"]},
    {"name": "Thales Italia",      "cluster": "Leader integrato",     "aliases": ["thales italia", "thales alenia", "thales group italy"]},
    {"name": "Accenture",          "cluster": "Leader integrato",     "aliases": ["accenture s.p.a", "accenture spa", "accenture italia"]},
    {"name": "ELT Group",          "cluster": "Specialista verticale","aliases": ["elt group", "elettronica s.p.a", "elettronica spa", "elettronica group"]},
    {"name": "Telsy",              "cluster": "Specialista verticale","aliases": ["telsy s.p.a", "telsy spa"]},
    {"name": "Cy4Gate",            "cluster": "Specialista verticale","aliases": ["cy4gate s.p.a", "cy4gate spa", "cy4gate"]},
    {"name": "IDS",                "cluster": "Specialista verticale","aliases": ["ids s.p.a", "ids spa", "ingegneria dei sistemi"]},
    {"name": "DEAS",               "cluster": "Specialista verticale","aliases": ["deas s.r.l", "deas srl", "deas s.p.a"]},
    {"name": "Exprivia",           "cluster": "Specialista verticale","aliases": ["exprivia s.p.a", "exprivia spa", "exprivia"]},
    {"name": "Engineering",        "cluster": "Generalista PA",       "aliases": ["engineering ingegneria", "engineering d.hub", "engineering s.p.a", "engineering spa"]},
    {"name": "IBM",                "cluster": "Generalista PA",       "aliases": ["ibm italia", "ibm s.p.a", "ibm spa"]},
    {"name": "AlmaViva",           "cluster": "Generalista PA",       "aliases": ["almaviva s.p.a", "almaviva spa", "almaviva the italian innovation company"]},
    {"name": "Kyndryl",            "cluster": "Generalista PA",       "aliases": ["kyndryl italia", "kyndryl s.p.a", "kyndryl spa"]},
    {"name": "CapGemini",          "cluster": "Generalista PA",       "aliases": ["capgemini italia", "cap gemini", "capgemini s.p.a", "capgemini spa"]},
    {"name": "Reply",              "cluster": "Generalista PA",       "aliases": ["reply s.p.a", "reply spa", "reply group"]},
    {"name": "Lutech",             "cluster": "Generalista PA",       "aliases": ["lutech s.p.a", "lutech spa", "lutech group"]},
    {"name": "Rheinmetall Italia", "cluster": "Generalista PA",       "aliases": ["rheinmetall italia", "rheinmetall italy"]},
    {"name": "BIP",                "cluster": "Generalista PA",       "aliases": ["bip s.p.a", "bip spa", "business integration partners"]},
    {"name": "PwC Advisory",       "cluster": "Generalista PA",       "aliases": ["pwc advisory", "pricewaterhousecoopers advisory", "pwc italia"]},
    {"name": "Vitrociset",         "cluster": "Entrante / marginale", "aliases": ["vitrociset s.p.a", "vitrociset spa"]},
    {"name": "BAE Systems Italy",  "cluster": "Entrante / marginale", "aliases": ["bae systems italy", "bae systems italia", "bae systems"]},
    {"name": "DXC Technology",     "cluster": "Entrante / marginale", "aliases": ["dxc technology", "dxc technology italia", "dxc"]},
    {"name": "Tinexta Cyber",      "cluster": "Entrante / marginale", "aliases": ["tinexta cyber", "tinexta s.p.a", "tinexta spa"]},
]

NOTA_METODOLOGICA_COMPETITOR = (
    "NOTA METODOLOGICA: La quota stimata si basa esclusivamente sulle gare pubbliche "
    "tracciabili (ANAC/BDNCP e fonti aperte) raccolte nel perimetro degli 11 player "
    "monitorati, in cui il competitor risulta aggiudicatario. Non rappresenta la quota "
    "di mercato assoluta: una parte rilevante delle forniture Difesa transita da "
    "procedure riservate, accordi quadro o canali con visibilita parziale. "
    "Utilizzare come indicatore relativo, non come dato certo."
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("defence_scraper")


def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_json(path, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError) as exc:
        logger.error("Impossibile leggere %s: %s. Uso il default.", path, exc)
        return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def load_history():
    return load_json(HISTORY_PATH, {})


def save_history(history):
    save_json(HISTORY_PATH, history)


def load_rotation_index():
    state = load_json(ROTATION_STATE_PATH, {"next_index": 0})
    idx = state.get("next_index", 0)
    if not isinstance(idx, int) or idx < 0 or idx >= len(PLAYERS):
        idx = 0
    return idx


def save_rotation_index(next_index):
    save_json(ROTATION_STATE_PATH, {
        "next_index": next_index,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })


def select_players_for_run(start_index):
    selected = []
    idx = start_index
    for _ in range(min(PLAYERS_PER_RUN, len(PLAYERS))):
        selected.append(idx)
        idx = (idx + 1) % len(PLAYERS)
    return selected, idx


def match_competitor(text, competitor):
    if not text:
        return False
    text_lower = text.lower().strip()
    if competitor["name"].lower() in text_lower or text_lower in competitor["name"].lower():
        return True
    for alias in competitor.get("aliases", []):
        if alias.lower() in text_lower or text_lower in alias.lower():
            return True
    return False


def extract_competitor_stats(history):
    all_gare = []
    for player_name, rec in history.items():
        for gara in rec.get("gare", []):
            all_gare.append({
                "player": player_name,
                "cig": gara.get("cig", ""),
                "oggetto": gara.get("oggetto", ""),
                "aggiudicatario": gara.get("aggiudicatario", ""),
                "importo": gara.get("importo", ""),
                "anno": gara.get("anno", ""),
            })
    results = []
    for comp in COMPETITORS:
        seen_cigs = set()
        gare_vinte = []
        for gara in all_gare:
            if match_competitor(gara.get("aggiudicatario", ""), comp):
                cig = str(gara.get("cig", "")).strip().upper()
                if cig and cig not in seen_cigs:
                    seen_cigs.add(cig)
                    gare_vinte.append(gara)
                elif not cig:
                    gare_vinte.append(gara)
        results.append({
            "name": comp["name"],
            "cluster": comp["cluster"],
            "n_gare": len(gare_vinte),
            "gare": gare_vinte,
            "top_cigs": ", ".join([g["cig"] for g in gare_vinte[:5] if g["cig"]]),
        })
    total = sum(r["n_gare"] for r in results)
    for r in results:
        r["quota_pct"] = round(r["n_gare"] / total * 100, 2) if total > 0 else 0.0
    results.sort(key=lambda x: x["quota_pct"], reverse=True)
    return results, total


SYSTEM_PROMPT = (
    "Sei un analista di mercato senior specializzato nel segmento Difesa "
    "italiano. Rispondi sempre in italiano. Usa la ricerca web per trovare "
    "dati aggiornati e verificabili, citando le fonti. Quando un dato non e "
    "reperibile, dichiaralo esplicitamente come 'N/D' senza inventare numeri."
)

PLAYER_PROMPT_TEMPLATE = (
    "Effettua una market analysis sul seguente player del segmento Difesa "
    "italiano per conto di NTT DATA Italia, che vuole aggredire questo "
    "mercato con un cluster Difesa unificato.\n\n"
    "PLAYER: {name}\n"
    "MERCATO: {market}\n"
    "COMPLESSITA PROCUREMENT: {complexity}\n\n"
    "Cerca attivamente sul web e raccogli i seguenti dati:\n"
    "1. Fatturato piu recente disponibile (con anno).\n"
    "2. Stima della spesa IT annua, con la fonte o la metodologia di stima.\n"
    "3. Servizi IT acquistati o di interesse (cloud, cybersecurity, system "
    "integration, software, ecc.).\n"
    "4. Piano strategico digitale o di trasformazione IT, se esiste.\n"
    "5. Storico delle gare/contratti rilevanti (preferibilmente da banca dati "
    "ANAC) con CIG, importo e aggiudicatario quando disponibili.\n"
    "6. Trend futuri attesi (investimenti, programmi, digitalizzazione).\n"
    "7. Note utili per NTT DATA (opportunita, punti di ingresso, rischi).\n\n"
    "Rispondi ESCLUSIVAMENTE con un oggetto JSON valido, senza testo prima o "
    "dopo, senza backtick e senza markdown. Struttura esatta:\n"
    "{{\n"
    '  "fatturato": "stringa",\n'
    '  "spesa_it_stimata": "stringa",\n'
    '  "fonte_stima": "stringa",\n'
    '  "servizi_it_acquistati": "stringa",\n'
    '  "piano_strategico_it": "stringa",\n'
    '  "trend_futuro": "stringa",\n'
    '  "note_ntt_data": "stringa",\n'
    '  "gare": [\n'
    '    {{"cig": "stringa", "oggetto": "stringa", "importo": "stringa", '
    '"aggiudicatario": "stringa", "anno": "stringa"}}\n'
    "  ]\n"
    "}}\n"
    "Se un campo non e reperibile usa il valore 'N/D'. Se non trovi gare usa "
    "una lista vuota []."
)


def extract_text_from_content(content_blocks):
    parts = []
    for block in content_blocks:
        if getattr(block, "type", None) == "text":
            parts.append(block.text)
    return "\n".join(parts).strip()


def parse_player_json(raw_text):
    text = raw_text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("Nessun oggetto JSON trovato nella risposta")
    return json.loads(text[start:end + 1])


def call_api_for_player(client, player):
    prompt = PLAYER_PROMPT_TEMPLATE.format(
        name=player["name"],
        market=player["market"],
        complexity=player["complexity"],
    )
    messages = [{"role": "user", "content": prompt}]
    tools = [{"type": "web_search_20250305", "name": "web_search", "max_uses": WEB_SEARCH_MAX_USES}]
    for _ in range(8):
        response = client.messages.create(
            model=MODEL, max_tokens=4096, system=SYSTEM_PROMPT,
            tools=tools, messages=messages,
        )
        if response.stop_reason in ("end_turn", "stop_sequence", "max_tokens"):
            return extract_text_from_content(response.content)
        if response.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": response.content})
            messages.append({"role": "user", "content": "Prosegui e fornisci la risposta finale in JSON."})
            continue
        return extract_text_from_content(response.content)
    return extract_text_from_content(response.content)


def classify_exception(exc):
    status = getattr(exc, "status_code", None)
    name = exc.__class__.__name__
    if status == 429 or "rate_limit" in str(exc).lower() or "RateLimit" in name:
        return "429 rate_limit_error"
    if "timeout" in str(exc).lower() or "Timeout" in name:
        return "timeout"
    if status == 404:
        return "404 model/endpoint not found"
    if status is not None:
        return "HTTP %s" % status
    return name


def is_rate_limit(exc):
    status = getattr(exc, "status_code", None)
    name = exc.__class__.__name__
    return status == 429 or "rate_limit" in str(exc).lower() or "RateLimit" in name


def fetch_player_data(client, player):
    attempt = 0
    while True:
        try:
            raw = call_api_for_player(client, player)
            try:
                data = parse_player_json(raw)
            except (ValueError, json.JSONDecodeError) as parse_exc:
                logger.error("Player '%s': JSON non parsato (%s). Risposta: %.200s", player["name"], parse_exc, raw)
                return None, "JSON non parsato"
            logger.info("Player '%s': SUCCESSO", player["name"])
            return data, "successo"
        except Exception as exc:
            kind = classify_exception(exc)
            if is_rate_limit(exc) and attempt < MAX_RETRIES:
                wait = BACKOFF_SECONDS[min(attempt, len(BACKOFF_SECONDS) - 1)]
                logger.error("Player '%s': %s (tentativo %d/%d). Attendo %ds.", player["name"], kind, attempt + 1, MAX_RETRIES, wait)
                time.sleep(wait)
                attempt += 1
                continue
            logger.error("Player '%s': ERRORE definitivo: %s", player["name"], kind)
            return None, kind


def build_record(player, data, status, timestamp):
    gare = data.get("gare", []) if isinstance(data, dict) else []
    if not isinstance(gare, list):
        gare = []
    return {
        "name": player["name"], "market": player["market"], "complexity": player["complexity"],
        "fatturato": (data or {}).get("fatturato", "N/D"),
        "spesa_it_stimata": (data or {}).get("spesa_it_stimata", "N/D"),
        "fonte_stima": (data or {}).get("fonte_stima", "N/D"),
        "servizi_it_acquistati": (data or {}).get("servizi_it_acquistati", "N/D"),
        "piano_strategico_it": (data or {}).get("piano_strategico_it", "N/D"),
        "trend_futuro": (data or {}).get("trend_futuro", "N/D"),
        "note_ntt_data": (data or {}).get("note_ntt_data", "N/D"),
        "gare": gare, "n_gare": len(gare), "status": status, "updated_at": timestamp,
    }


def compute_delta(old_record, new_record):
    if old_record is None:
        return "Primo rilevamento"
    changes = []
    for field, label in [("spesa_it_stimata", "Spesa IT"), ("fatturato", "Fatturato"), ("trend_futuro", "Trend")]:
        if str(old_record.get(field, "N/D")) != str(new_record.get(field, "N/D")):
            changes.append(label)
    diff = new_record.get("n_gare", 0) - old_record.get("n_gare", 0)
    if diff != 0:
        changes.append("Gare (%s%d)" % ("+" if diff > 0 else "", diff))
    return ("Variazioni: " + ", ".join(changes)) if changes else "Nessuna variazione"


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1F3864", bold=True, size=14)
COMP_HEADER_FILL = PatternFill("solid", fgColor="0A2E5C")
COMP_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="E8F0FE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row_idx, n_cols, fill=None, font=None):
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def autosize_columns(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_excel(records_by_name, deltas_by_name, run_date, competitor_stats, total_gare_competitor):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Dashboard"
    ws.cell(row=1, column=1, value="Market Analysis Difesa Italia - NTT DATA").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Aggiornato al %s" % run_date).font = Font(italic=True, color="595959")
    headers = ["Player", "Mercato", "Complessita Procurement", "Fatturato", "Spesa IT Stimata",
               "Fonte Stima", "Servizi IT Acquistati", "Piano Strategico IT", "N. Gare Trovate",
               "Trend Futuro", "Delta vs Ieri", "Note per NTT DATA"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=head)
    style_header_row(ws, 4, len(headers))
    row = 5
    for player in PLAYERS:
        rec = records_by_name.get(player["name"]) or build_record(player, {}, "nessun dato storico", run_date)
        values = [rec["name"], rec["market"], rec["complexity"], rec["fatturato"],
                  rec["spesa_it_stimata"], rec["fonte_stima"], rec["servizi_it_acquistati"],
                  rec["piano_strategico_it"], rec["n_gare"], rec["trend_futuro"],
                  deltas_by_name.get(player["name"], "N/D"), rec["note_ntt_data"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
        row += 1
    autosize_columns(ws, [26, 12, 14, 22, 22, 24, 30, 30, 12, 30, 22, 36])
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("Storico Gare")
    ws2.cell(row=1, column=1, value="Storico Gare ANAC e contratti rilevanti").font = TITLE_FONT
    gare_headers = ["Player", "CIG", "Oggetto", "Importo", "Aggiudicatario", "Anno"]
    for col, head in enumerate(gare_headers, start=1):
        ws2.cell(row=3, column=col, value=head)
    style_header_row(ws2, 3, len(gare_headers))
    grow = 4
    any_gara = False
    for player in PLAYERS:
        rec = records_by_name.get(player["name"])
        if not rec:
            continue
        for gara in rec.get("gare", []):
            any_gara = True
            for col, val in enumerate([rec["name"], gara.get("cig", "N/D"), gara.get("oggetto", "N/D"),
                                        gara.get("importo", "N/D"), gara.get("aggiudicatario", "N/D"),
                                        gara.get("anno", "N/D")], start=1):
                cell = ws2.cell(row=grow, column=col, value=val)
                cell.alignment = WRAP
                cell.border = BORDER
            grow += 1
    if not any_gara:
        ws2.cell(row=4, column=1, value="Nessuna gara trovata nei dati disponibili.")
    autosize_columns(ws2, [26, 20, 44, 18, 30, 8])
    ws2.freeze_panes = "A4"

    ws3 = wb.create_sheet("Dettaglio Player")
    drow = 1
    field_labels = [("market", "Mercato"), ("complexity", "Complessita Procurement"),
                    ("fatturato", "Fatturato"), ("spesa_it_stimata", "Spesa IT Stimata"),
                    ("fonte_stima", "Fonte Stima"), ("servizi_it_acquistati", "Servizi IT Acquistati"),
                    ("piano_strategico_it", "Piano Strategico IT"), ("trend_futuro", "Trend Futuro"),
                    ("note_ntt_data", "Note per NTT DATA"), ("n_gare", "N. Gare Trovate"),
                    ("status", "Esito ultimo rilevamento"), ("updated_at", "Ultimo aggiornamento")]
    for player in PLAYERS:
        rec = records_by_name.get(player["name"]) or build_record(player, {}, "nessun dato storico", run_date)
        title_cell = ws3.cell(row=drow, column=1, value=rec["name"])
        title_cell.font = TITLE_FONT
        ws3.merge_cells(start_row=drow, start_column=1, end_row=drow, end_column=2)
        drow += 1
        for key, label in field_labels:
            lc = ws3.cell(row=drow, column=1, value=label)
            lc.font = Font(bold=True)
            lc.alignment = Alignment(vertical="top")
            lc.border = BORDER
            vc = ws3.cell(row=drow, column=2, value=rec.get(key, "N/D"))
            vc.alignment = WRAP
            vc.border = BORDER
            drow += 1
        drow += 1
    autosize_columns(ws3, [28, 70])

    ws4 = wb.create_sheet("Analisi Competitor")
    title_cell = ws4.cell(row=1, column=1,
        value="Analisi Competitor - Quota Stimata Mercato Difesa IT | NTT DATA Italia | " + run_date)
    title_cell.font = TITLE_FONT
    ws4.merge_cells("A1:G1")
    ws4["A1"].alignment = Alignment(horizontal="center")

    comp_headers = ["Competitor", "Cluster", "N. Gare Vinte", "Quota Stimata %",
                    "Gare Principali (CIG)", "Aggiornato al", "Note"]
    for col, head in enumerate(comp_headers, start=1):
        cell = ws4.cell(row=3, column=col, value=head)
        cell.fill = COMP_HEADER_FILL
        cell.font = COMP_HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    autosize_columns(ws4, [22, 22, 16, 16, 45, 16, 35])

    data_start = 4
    for row_idx, comp in enumerate(competitor_stats, start=data_start):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        values = [comp["name"], comp["cluster"], comp["n_gare"], comp["quota_pct"] / 100,
                  comp.get("top_cigs", ""), run_date,
                  "Dato da gare tracciate nel perimetro dei player monitorati"]
        for col_idx, val in enumerate(values, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 4:
                cell.number_format = "0.00%"
    data_end = data_start + len(competitor_stats) - 1

    nota_row = data_end + 2
    ws4.merge_cells("A%d:G%d" % (nota_row, nota_row))
    nota_cell = ws4["A%d" % nota_row]
    nota_cell.value = NOTA_METODOLOGICA_COMPETITOR
    nota_cell.font = Font(italic=True, size=9, color="666666")
    nota_cell.alignment = Alignment(wrap_text=True)
    ws4.row_dimensions[nota_row].height = 50

    if competitor_stats:
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
        chart.title = "Quota Stimata % per Competitor (gare Difesa pubblicamente tracciate)"
        chart.y_axis.title = "Competitor"
        chart.x_axis.title = "Quota Stimata %"
        chart.style = 10
        chart.width = 28
        chart.height = 18
        data_ref = Reference(ws4, min_col=4, min_row=data_start, max_row=data_end)
        cats_ref = Reference(ws4, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        ws4.add_chart(chart, "A%d" % (nota_row + 3))
    ws4.freeze_panes = "A4"

    filename = "market_analysis_difesa_%s.xlsx" % run_date
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    logger.info("Excel generato: %s", path)
    return path


def build_email_html(records_by_name, deltas_by_name, run_date, kpis, competitor_stats):
    rows_html = []
    for player in PLAYERS:
        rec = records_by_name.get(player["name"]) or build_record(player, {}, "nessun dato storico", run_date)
        rows_html.append(
            "<tr><td style='padding:8px;border:1px solid #ddd;'>%s</td>"
            "<td style='padding:8px;border:1px solid #ddd;'>%s</td>"
            "<td style='padding:8px;border:1px solid #ddd;'>%s</td>"
            "<td style='padding:8px;border:1px solid #ddd;text-align:center;'>%s</td>"
            "<td style='padding:8px;border:1px solid #ddd;'>%s</td></tr>" % (
                rec["name"], rec["market"], rec["spesa_it_stimata"],
                rec["n_gare"], deltas_by_name.get(player["name"], "N/D"),
            )
        )

    kpi_box = ("<div style='display:inline-block;width:23%%;min-width:140px;margin:6px;"
               "padding:16px;background:#1F3864;color:#fff;border-radius:8px;"
               "text-align:center;vertical-align:top;'>"
               "<div style='font-size:28px;font-weight:bold;'>%s</div>"
               "<div style='font-size:12px;opacity:0.9;'>%s</div></div>")
    kpi_html = "".join(kpi_box % (value, label) for label, value in kpis)

    top5 = competitor_stats[:5]
    comp_rows = "".join(
        "<tr><td style='padding:8px;border:1px solid #ddd;'>%s</td>"
        "<td style='padding:8px;border:1px solid #ddd;'>%s</td>"
        "<td style='padding:8px;border:1px solid #ddd;text-align:center;font-weight:bold;'>%d</td>"
        "<td style='padding:8px;border:1px solid #ddd;text-align:center;font-weight:bold;color:#1F3864;'>%.2f%%</td></tr>"
        % (c["name"], c["cluster"], c["n_gare"], c["quota_pct"]) for c in top5
    )
    comp_section = (
        "<h3 style='color:#1F3864;margin-top:28px;'>Top 5 Competitor per Quota Stimata</h3>"
        "<table style='border-collapse:collapse;width:100%%;font-size:13px;'>"
        "<thead><tr style='background:#0A2E5C;color:#fff;'>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Competitor</th>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Cluster</th>"
        "<th style='padding:8px;border:1px solid #ddd;'>N. Gare</th>"
        "<th style='padding:8px;border:1px solid #ddd;'>Quota Stimata %%</th>"
        "</tr></thead><tbody>%s</tbody></table>"
        "<p style='font-size:11px;color:#888;font-style:italic;'>%s</p>"
    ) % (comp_rows, NOTA_METODOLOGICA_COMPETITOR) if top5 else ""

    return (
        "<html><body style='font-family:Arial,sans-serif;color:#222;max-width:800px;margin:auto;padding:20px;'>"
        "<h2 style='color:#1F3864;'>Market Analysis Difesa Italia - NTT DATA</h2>"
        "<p>Report automatico del <strong>%s</strong>. In allegato l'Excel completo "
        "(Dashboard, Storico Gare, Dettaglio Player, Analisi Competitor).</p>"
        "<div style='text-align:center;margin:16px 0;'>%s</div>"
        "<h3 style='color:#1F3864;'>Quadro player</h3>"
        "<table style='border-collapse:collapse;width:100%%;font-size:13px;'>"
        "<thead><tr style='background:#1F3864;color:#fff;'>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Player</th>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Mercato</th>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Spesa IT Stimata</th>"
        "<th style='padding:8px;border:1px solid #ddd;'>N. Gare</th>"
        "<th style='padding:8px;border:1px solid #ddd;text-align:left;'>Delta vs Ieri</th>"
        "</tr></thead><tbody>%s</tbody></table>%s"
        "<p style='font-size:11px;color:#888;margin-top:20px;'>Scaglionamento a "
        "rotazione attivo: %d player processati per run.</p>"
        "</body></html>"
    ) % (run_date, kpi_html, "".join(rows_html), comp_section, PLAYERS_PER_RUN)


def send_email(html_body, attachment_path, run_date):
    host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ["SMTP_USER"]
    password = os.environ["SMTP_PASS"]
    recipient = os.environ.get("RECIPIENT_EMAIL", "jacopo.roccella@nttdata.com")
    msg = MIMEMultipart()
    msg["From"] = user
    msg["To"] = recipient
    msg["Subject"] = "Market Analysis Difesa Italia - %s" % run_date
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    with open(attachment_path, "rb") as fh:
        part = MIMEApplication(fh.read(), _subtype="xlsx")
    part.add_header("Content-Disposition", "attachment", filename=os.path.basename(attachment_path))
    msg.attach(part)
    logger.info("Invio mail a %s via %s:%d", recipient, host, port)
    with smtplib.SMTP(host, port) as server:
        server.starttls()
        server.login(user, password)
        server.sendmail(user, [recipient], msg.as_string())
    logger.info("Mail inviata correttamente.")


def main():
    ensure_dirs()
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    timestamp = datetime.now(timezone.utc).isoformat()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        logger.error("ANTHROPIC_API_KEY mancante. Impossibile proseguire.")
        sys.exit(1)

    client = Anthropic(api_key=api_key)
    history = load_history()
    start_index = load_rotation_index()
    selected_indices, next_index = select_players_for_run(start_index)
    selected_names = [PLAYERS[i]["name"] for i in selected_indices]
    logger.info("Run del %s. Processo %d player: %s", run_date, len(selected_indices), ", ".join(selected_names))

    fresh_success = 0
    errors = 0
    for pos, idx in enumerate(selected_indices):
        player = PLAYERS[idx]
        logger.info("Interrogo player %d/%d: %s", pos + 1, len(selected_indices), player["name"])
        data, status = fetch_player_data(client, player)
        if data is not None:
            new_record = build_record(player, data, status, timestamp)
            new_record["delta"] = compute_delta(history.get(player["name"]), new_record)
            history[player["name"]] = new_record
            fresh_success += 1
        else:
            errors += 1
            logger.error("Player '%s' non aggiornato (%s). Mantengo storico.", player["name"], status)
        if pos < len(selected_indices) - 1:
            logger.info("Pausa di %d secondi.", PAUSE_BETWEEN_PLAYERS)
            time.sleep(PAUSE_BETWEEN_PLAYERS)

    save_rotation_index(next_index)
    save_history(history)

    records_by_name = {}
    deltas_by_name = {}
    players_with_it_estimate = 0
    total_gare = 0
    total_changes = 0
    for player in PLAYERS:
        rec = history.get(player["name"])
        if rec is None:
            continue
        records_by_name[player["name"]] = rec
        delta = rec.get("delta", "Dato storico (non aggiornato oggi)")
        deltas_by_name[player["name"]] = delta
        if rec.get("spesa_it_stimata", "N/D") not in ("N/D", "", None):
            players_with_it_estimate += 1
        total_gare += rec.get("n_gare", 0)
        if delta.startswith("Variazioni"):
            total_changes += 1

    if not records_by_name:
        logger.error("ZERO player con dati validi. Non invio la mail. Esco con codice 1.")
        sys.exit(1)

    competitor_stats, total_gare_competitor = extract_competitor_stats(history)
    comp_con_gare = sum(1 for c in competitor_stats if c["n_gare"] > 0)
    logger.info("Analisi competitor: %d su %d con almeno una gara su %d totali.",
                comp_con_gare, len(COMPETITORS), total_gare_competitor)

    kpis = [
        ("Player monitorati", len(records_by_name)),
        ("Con stima spesa IT", players_with_it_estimate),
        ("Gare trovate", total_gare),
        ("Variazioni vs ieri", total_changes),
    ]

    excel_path = build_excel(records_by_name, deltas_by_name, run_date, competitor_stats, total_gare_competitor)
    html_body = build_email_html(records_by_name, deltas_by_name, run_date, kpis, competitor_stats)

    try:
        send_email(html_body, excel_path, run_date)
    except Exception as exc:
        logger.error("Invio mail fallito: %s", exc)
        sys.exit(1)

    logger.info("Run completato. Player freschi: %d, errori: %d, totale in report: %d.",
                fresh_success, errors, len(records_by_name))


if __name__ == "__main__":
    main()
